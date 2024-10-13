import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

archivo_xls = "informe_gastos.xlsx"
hoja = "Gastos"

def validar_fecha(fecha):
    try:
        return datetime.strptime(fecha, "%d-%m-%Y")
    except ValueError:
        return None

def gastos_dat():
    gastos = []
    
    while True:
        print("Ingrese sus gastos o escriba 'fin' para terminar.")
        
        fecha = input("Fecha en formato (DD-MM-YYYY): ")
        if fecha.lower() == 'fin':
            break
        
        fecha_valida = validar_fecha(fecha)
        if not fecha_valida:
            print("Formato de fecha incorrecto el correcto es DD-MM-YYYY.")
            continue
        
        try:
            cantidad = input("Cantidad: ")
            if cantidad.lower() == 'fin':
                break
            cantidad = float(cantidad)
            if cantidad < 0:
                print("El gasto no puede ser negativo.")
                continue
        except ValueError:
            print("Por favor ingrese un número válido.")
            continue
        
        descripcion = input("Descripción: ")
        if descripcion.lower() == 'fin':
            break
        
        gastos.append({
            "Fecha": fecha,
            "Cantidad": cantidad,
            "Descripción": descripcion,
        })
        
        print("Gasto agregado correctamente.\n")
    
    return gastos

def crear_archivo_si_no_existe():
    if not os.path.exists(archivo_xls):
        print(f"El archivo '{archivo_xls}' no existe. Creando uno nuevo...")

        libro_trabajo = Workbook()
        hoja_trabajo = libro_trabajo.active
        hoja_trabajo.title = hoja
        hoja_trabajo.append(["Fecha", "Cantidad", "Descripción"])

        libro_trabajo.save(archivo_xls)
        print(f"Archivo '{archivo_xls}' creado correctamente.")
    else:
        print(f"El archivo '{archivo_xls}' ya existe. Se usará para guardar los datos.")

def ingresar_datos(gastos):
    crear_archivo_si_no_existe()
    libro_trabajo = openpyxl.load_workbook(archivo_xls)
    
    if hoja in libro_trabajo.sheetnames:
        hoja_trabajo = libro_trabajo[hoja]
    else:

        hoja_trabajo = libro_trabajo.create_sheet(hoja)
        hoja_trabajo.append(["Fecha", "Cantidad", "Descripción"])
    
    for gasto in gastos:
        hoja_trabajo.append([gasto["Fecha"], gasto["Cantidad"], gasto["Descripcion"]])
    
    libro_trabajo.save(archivo_xls)
    print(f"Los gastos se guardaron correctamente en '{archivo_xls}'.")

def resumen(gastos):
    if not gastos:
        print("No hay gastos para mostrar.")
        return
    
    total_gastos = sum(gasto["Cantidad"] for gasto in gastos)
    cantidad_de_gastos = len(gastos)
    maximo = max(gastos, key=lambda x: x["Cantidad"])
    minimo = min(gastos, key=lambda x: x["Cantidad"])
    
    print("\nResumen de gastos:")
    print(f"Total de gastos: {total_gastos:.2f}")
    print(f"Cantidad de gastos: {cantidad_de_gastos}")
    print(f"Gasto más alto: {maximo['Descripcion']} - {maximo['Cantidad']:.2f} ({maximo['Fecha']})")
    print(f"Gasto más bajo: {minimo['Descripcion']} - {minimo['Cantidad']:.2f} ({minimo['Fecha']})")

def start():
    print("Gestión de gastos")
    
    gastos = gastos_dat()
    
    if gastos:
        ingresar_datos(gastos)
        resumen(gastos)
    else:
        print("No se ingresaron gastos.")
    
    print(f"Los gastos se guardaron en '{archivo_xls}'.")

start()